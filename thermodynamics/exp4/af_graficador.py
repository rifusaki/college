# -*- coding: utf-8 -*-
"""
Created on Mon Sep 11 21:53:24 2023

@author: afsuarez
"""

import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit

wb=load_workbook("D:/Informacion/Desktop/termo/termo frio.xlsx") #Archivo de donde salen los datos
ws=wb["Hoja1"]                           #hoja donde estan los datos
T1=[]
t1=[]                                 #Lista que sera llenada de datos
#X=np.arange(295,350,9 )
for i in np.arange(5,45,1):                         #de donde a donde estan los datos
  T1.append(ws[f"F{i}"].value)
  t1.append(ws[f"G{i}"].value)

######REGRESIÓN
def f(x,Ti,Tf,t):
     return Tf+np.multiply((Ti-Tf),np.exp(-x/t))
arg,cov=curve_fit(f,t1,T1)
x=[i for i in range(0,820,20)]
print(arg)
# param,cov=curve_fit(f,T,R)
#Crear una figura y un conjunto de ejes#Graficar los datos
fig, ax = plt.subplots(figsize=(8,6))
ax.set(xlabel="Tiempo(s) ",ylabel=r"Temperatura(c°)",title=" Temperatura vs tiempo ")
plt.plot(t1,T1,"o",linewidth=1,color="blue",markersize=4,alpha=1,label="Datos experimentales")
plt.plot(t1,f(np.array(t1), 67.5628304   ,26.81808506, 294.7435808),"--",linewidth=1,color="blue",markersize=4,alpha=0.5,label=r"$T(t)=26.81+(67.56 − 26.81 )e^{−t/294.74} $")
#plt.plot(t1,Tt)
plt.legend()
ax.grid()
#plt.savefig("3Tvst.png",dpi=300)
plt.show()
